import logging
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from flask import Blueprint
from flask import render_template, url_for, flash, redirect, request, abort, make_response
from onetouch import db, bcrypt
from onetouch.models import Student, User, StudentDebt, School, TransactionRecord
from onetouch.students.forms import EditStudentModalForm, RegisterStudentModalForm
from flask_login import login_required, current_user
from flask import jsonify
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

from onetouch.students.functons import check_if_plus_one, get_school_year_for_change
from onetouch.suppliers.functions import convert_to_latin
from sqlalchemy.exc import SQLAlchemyError
import logging
logger = logging.getLogger(__name__)

students = Blueprint('students', __name__)


# Ova funkcija će proveriti da li je korisnik ulogovan pre nego što pristupi zaštićenoj ruti
@students.before_request
def require_login():
    if request.endpoint and not current_user.is_authenticated:
        return redirect(url_for('users.login'))


@students.route('/student_list', methods=['GET', 'POST'])
def student_list():
    try:
        route_name = request.endpoint
        school = School.query.first()
        danas = datetime.now()
        plus_one_button = check_if_plus_one(school)
        active_date_start = danas.replace(month=8, day=15)
        active_date_end = danas.replace(month=9, day=30)
        
        # Provera isteka licence
        license_expired = False
        if school and school.license_expiry_date:
            days_left = school.days_until_license_expiry()
            if days_left is not None and days_left <= 0:
                license_expired = True
        
        register_form = RegisterStudentModalForm()
        edit_form = EditStudentModalForm()
        
        if edit_form.validate_on_submit() and request.form.get('submit_edit'):
            try:
                student = Student.query.get(request.form.get('student_id'))
                logging.debug(f'debug- edit mode: {student.student_name=} {student.student_surname=}')
                logging.debug(f'{edit_form.student_name.data=} {edit_form.student_surname.data=} {edit_form.send_mail.data=}')
                logging.debug(f'{request.form.get("send_mail")=}')
                student.student_name = convert_to_latin(edit_form.student_name.data.strip().title())
                student.student_surname = convert_to_latin(" ".join(word.title() for word in edit_form.student_surname.data.strip().split()))
                student.student_class = edit_form.student_class.data
                student.student_section = edit_form.student_section.data
                student.parent_email = edit_form.parent_email.data
                student.send_mail = edit_form.send_mail.data
                student.print_payment = edit_form.print_payment.data
                db.session.commit()
                flash(f'Podaci o učeniku "{student.student_name} {student.student_surname}" su ažurirani.', 'success')
                return redirect(url_for('students.student_list'))
            except SQLAlchemyError as e:
                db.session.rollback()
                logger.error(f"Greška pri ažuriranju učenika: {str(e)}")
                flash('Došlo je do greške pri ažuriranju podataka o učeniku.', 'danger')
                return render_template('errors/500.html'), 500
                
        if register_form.validate_on_submit() and request.form.get('submit_register'):
            try:
                student = Student(
                                student_name=convert_to_latin(register_form.student_name.data.strip().title()),
                                student_surname=convert_to_latin(" ".join(word.title() for word in register_form.student_surname.data.strip().split())),
                                student_class=str(register_form.student_class.data),
                                student_section=register_form.student_section.data,
                                parent_email=register_form.parent_email.data,
                                send_mail=0,
                                print_payment=0
                            )

                logging.debug(student.student_name, student.student_surname, student.student_class)
                db.session.add(student)
                db.session.commit()
                logging.debug("inputi su validni")
                flash(f'Učenik "{student.student_name} {student.student_surname}" je registrovan.', 'success')
                return redirect(url_for('students.student_list'))
            except SQLAlchemyError as e:
                db.session.rollback()
                logger.error(f"Greška pri registraciji učenika: {str(e)}")
                flash('Došlo je do greške pri registraciji učenika.', 'danger')
                return render_template('errors/500.html'), 500
                
        return render_template('student_list.html', 
                                title='Učenici', 
                                legend="Učenici",
                                register_form=register_form, 
                                edit_form=edit_form,
                                plus_one_button=plus_one_button,
                                active_date_start=active_date_start,
                                active_date_end=active_date_end,
                                danas=danas,
                                route_name=route_name,
                                license_expired=license_expired,
                                school=school)
    except Exception as e:
        logger.error(f"Neočekivana greška u student_list: {str(e)}")
        flash('Došlo je do neočekivane greške.', 'danger')
        return render_template('errors/500.html'), 500


@students.route('/api/students_list')
def api_students_list():
    try:
        students_query = Student.query.filter(Student.student_class < 9)
        
        # search filter
        search = request.args.get('search[value]')
        student_class = request.args.get('searchRazred')
        student_section = request.args.get('searchOdeljenje')
        logging.debug(f'debug: {search=} {student_class=} {student_section=}')
        
        try:
            if search:
                students_query = students_query.filter(db.or_(
                    Student.student_name.like(f'%{search}%'),
                    Student.student_surname.like(f'%{search}%'),
                    Student.id.like(f'%{search}%'))
                )
            if student_class:
                students_query = students_query.filter(Student.student_class == student_class)
            if student_section:
                students_query = students_query.filter(Student.student_section == student_section)
            total_filtered = students_query.count()
        except SQLAlchemyError as e:
            logger.error(f"Greška pri filtriranju učenika: {str(e)}")
            return jsonify({
                'error': 'Došlo je do greške pri pretraživanju učenika',
                'data': [],
                'recordsFiltered': 0,
                'recordsTotal': 0,
                'draw': request.args.get('draw', type=int)
            }), 500
        
        # sorting
        order = []
        i = 0
        while True:
            col_index = request.args.get(f'order[{i}][column]')
            if col_index is None:
                break
            col_name = request.args.get(f'columns[{col_index}][data]')
            if col_name not in ['id', 'student_name', 'student_surname', 'student_class', 'student_section', 'parent_email']:
                col_name = 'parent_email'
            descending = request.args.get(f'order[{i}][dir]') == 'desc'
            col = getattr(Student, col_name)
            if descending:
                col = col.desc()
            order.append(col)
            i += 1
        if order:
            students_query = students_query.order_by(*order)
        
        # pagination
        start = request.args.get('start', type=int)
        length = request.args.get('length', type=int)
        
        try:
            students_query = students_query.offset(start).limit(length)
            students_list = []
            for student in students_query:
                new_dict = {
                    'id': str(student.id).zfill(4),
                    'student_name': student.student_name,
                    'student_surname': student.student_surname,
                    'student_class': student.student_class,
                    'student_section': student.student_section,
                    'parent_email': student.parent_email if student.parent_email else "",
                    'send_mail': student.send_mail,
                    'print_payment': student.print_payment
                }
                students_list.append(new_dict)
            logging.debug(f'{students_list=}')
            
            return {
                'data': students_list,
                'recordsFiltered': total_filtered,
                'recordsTotal': total_filtered,
                'draw': request.args.get('draw', type=int),
            }
        except SQLAlchemyError as e:
            logger.error(f"Greška pri učitavanju učenika: {str(e)}")
            return jsonify({
                'error': 'Došlo je do greške pri učitavanju učenika',
                'data': [],
                'recordsFiltered': 0,
                'recordsTotal': 0,
                'draw': request.args.get('draw', type=int)
            }), 500
            
    except Exception as e:
        logger.error(f"Neočekivana greška u api_students_list: {str(e)}")
        return jsonify({
            'error': 'Došlo je do neočekivane greške',
            'data': [],
            'recordsFiltered': 0,
            'recordsTotal': 0,
            'draw': request.args.get('draw', type=int)
        }), 500



@students.route('/class_plus_one', methods=['GET', 'POST'])
def class_plus_one():
    try:
        # Dohvatanje podataka iz baze
        try:
            students = Student.query.filter(Student.student_class < 100).all()
            school = School.query.first()
            last_class_plus_one = school.class_plus_one
        except SQLAlchemyError as e:
            logger.error(f"Greška pri dohvatanju podataka iz baze: {str(e)}")
            flash('Došlo je do greške pri učitavanju podataka.', 'danger')
            return redirect(url_for('students.student_list'))
        
        if not students:
            logger.warning("Nema učenika za prebacivanje u sledeći razred")
            flash('Nema učenika kojima treba promeniti razred.', 'info')
            return redirect(url_for('students.student_list'))
        
        # trenutni datum
        now = datetime.now().date()
        
        # Proveravamo da li smo u dozvoljenom periodu (15. avgust - 30. septembar)
        if not ((now.month == 8 and now.day >= 15) or now.month == 9):
            logger.warning(f'Pokušaj promene razreda van dozvoljenog perioda: {now}')
            flash('Promena razreda je moguća samo u periodu od 15. avgusta do 30. septembra.', 'warning')
            return redirect(url_for('students.student_list'))
        
        # Određujemo za koju školsku godinu bi bila trenutna promena
        current_change_school_year = get_school_year_for_change(now)
        logger.debug(f'{current_change_school_year=}')
        # Proveravamo da li je već rađena promena za istu školsku godinu
        if last_class_plus_one:
            last_change_school_year = get_school_year_for_change(last_class_plus_one)
            logger.debug(f'{last_change_school_year=}')
            
            if current_change_school_year == last_change_school_year:
                logger.info(f'Već je urađena promena za školsku godinu {current_change_school_year[0]}/{current_change_school_year[1]}. '
                            f'Poslednja promena: {last_class_plus_one}, trenutni datum: {now}')
                flash(f'Razred je već promenjen za školsku godinu {current_change_school_year[0]}/{current_change_school_year[1]}.', 'info')
                return redirect(url_for('students.student_list'))
        
        # Izvršavamo promenu razreda
        logger.info(f'Prebacivanje učenika u sledeći razred. Datum: {now}, '
                    f'Školska godina: {current_change_school_year[0]}/{current_change_school_year[1]}')
        
        try:
            for student in students:
                student.student_class = int(student.student_class) + 1
            school.class_plus_one = now
            db.session.commit()
            
            flash(f'Razred je uspešno promenjen za sve učenike.\nZa učenike koji su ponavljali razred, potrebno je ručno ažurirati njihove podatke.', 'success')
            logger.info(f'Uspešno promenjen razred za {len(students)} učenika.')
            
        except SQLAlchemyError as e:
            db.session.rollback()
            logger.error(f"Greška pri ažuriranju razreda: {str(e)}")
            flash('Došlo je do greške pri promeni razreda. Promene nisu sačuvane.', 'danger')
        
        return redirect(url_for('students.student_list'))
        
    except Exception as e:
        logger.error(f"Neočekivana greška u class_plus_one: {str(e)}")
        flash('Došlo je do neočekivane greške.', 'danger')
        return redirect(url_for('students.student_list'))


@students.route('/student/<int:student_id>/delete', methods=['POST'])
@login_required
def delete_student(student_id):
    try:
        # Provera da li student postoji
        try:
            student = Student.query.get(student_id)
            if not student:
                flash('Učenik nije pronađen.', 'danger')
                return redirect(url_for('students.student_list'))
        except SQLAlchemyError as e:
            logger.error(f"Greška pri dohvatanju učenika: {str(e)}")
            flash('Došlo je do greške pri pristupu podacima učenika.', 'danger')
            return redirect(url_for('students.student_list'))

        # Provera autentifikacije i lozinke
        if not current_user.is_authenticated:
            flash('Morate biti prijavljeni da biste pristupili ovoj stranici', 'danger')
            return redirect(url_for('users.login'))
        elif not bcrypt.check_password_hash(current_user.user_password, request.form.get("input_password")):
            logger.warning(f'Pogrešna lozinka pri pokušaju brisanja učenika: {student.student_name} {student.student_surname}')
            flash(f'Pogrešna lozinka! nije obrisan profil učenika: {student.student_name} {student.student_surname}', 'danger')
            return redirect(url_for('students.student_list'))
        
        # Provera zaduženja i brisanje/arhiviranje
        try:
            list_of_students_with_debts = TransactionRecord.query.filter_by(student_id=student.id).all()
            logger.debug(f'{list_of_students_with_debts=}')
            
            if list_of_students_with_debts:
                logger.info(f'Učenik {student.student_name} {student.student_surname} ima zaduženje, prebacuje se u arhivu (razred+100)')
                student.student_class += 100
            else:
                logger.info(f'Učenik {student.student_name} {student.student_surname} nema zaduženje, briše se iz baze')
                db.session.delete(student)
                
            db.session.commit()
            flash(f'Profil učenika "{student.student_name} {student.student_surname}" je obrisan.', 'success')
            
        except SQLAlchemyError as e:
            db.session.rollback()
            logger.error(f"Greška pri brisanju/arhiviranju učenika: {str(e)}")
            flash('Došlo je do greške pri brisanju učenika. Promene nisu sačuvane.', 'danger')
            
        return redirect(url_for("students.student_list"))
        
    except Exception as e:
        logger.error(f"Neočekivana greška u delete_student: {str(e)}")
        flash('Došlo je do neočekivane greške.', 'danger')
        return redirect(url_for('students.student_list'))


@students.route('/export_students_excel', methods=['GET'])
@login_required
def export_students_excel():
    try:
        # Dohvatanje parametara filtera iz URL-a
        student_class = request.args.get('razred', '')
        student_section = request.args.get('odeljenje', '')
        
        # Kreiranje početnog upita
        students_query = Student.query.filter(Student.student_class < 9)
        
        # Primena filtera
        if student_class:
            students_query = students_query.filter(Student.student_class == student_class)
        if student_section:
            students_query = students_query.filter(Student.student_section == student_section)
        
        # Sortiranje rezultata
        students_query = students_query.order_by(
            Student.student_class, 
            Student.student_section, 
            Student.student_surname, 
            Student.student_name
        )
        
        # Kreiranje novog Excel fajla
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Učenici"
        
        # Definisanje stilova
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        title_font = Font(name='Arial', size=12, bold=True)
        
        # Definisanje ivica
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Dodavanje informacija o filterima na vrh
        current_row = 1
        
        # Naslov izveštaja
        cell = ws.cell(row=current_row, column=1)
        cell.value = "Spisak učenika"
        cell.font = Font(name='Arial', size=14, bold=True)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        cell.alignment = Alignment(horizontal='center')
        current_row += 2  # Preskačemo jedan red
        
        # Informacije o filterima
        if student_class or student_section:
            filter_text = "Primenjeni filteri: "
            if student_class:
                filter_text += f"Razred: {student_class}"
            if student_class and student_section:
                filter_text += ", "
            if student_section:
                filter_text += f"Odeljenje: {student_section}"
                
            cell = ws.cell(row=current_row, column=1)
            cell.value = filter_text
            cell.font = title_font
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
            current_row += 2  # Preskačemo jedan red za razmak
        
        # Dodavanje zaglavlja
        headers = ['ID učenika', 'Ime', 'Prezime', 'Razred', 'Odeljenje', 'Mejl roditelja']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Podešavanje širine kolona
        column_widths = [15, 20, 20, 10, 15, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            
        # Pamtimo red zaglavlja za kasnije dodavanje podataka
        header_row = current_row
        
        # Dodavanje podataka (počinjemo sa sledećim redom nakon zaglavlja)
        for row_num, student in enumerate(students_query, 1):
            data_row = header_row + row_num
            
            # ID učenika
            cell = ws.cell(row=data_row, column=1)
            cell.value = str(student.id).zfill(4)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            
            # Ime
            cell = ws.cell(row=data_row, column=2)
            cell.value = student.student_name
            cell.border = thin_border
            
            # Prezime
            cell = ws.cell(row=data_row, column=3)
            cell.value = student.student_surname
            cell.border = thin_border
            
            # Razred
            cell = ws.cell(row=data_row, column=4)
            cell.value = student.student_class
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            
            # Odeljenje
            cell = ws.cell(row=data_row, column=5)
            cell.value = student.student_section
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            
            # Mejl roditelja
            cell = ws.cell(row=data_row, column=6)
            cell.value = student.parent_email if student.parent_email else ""
            cell.border = thin_border
        
        # Generisanje naziva fajla
        filename = "Spisak_ucenika"
        if student_class:
            filename += f"_razred_{student_class}"
        if student_section:
            filename += f"_odeljenje_{student_section}"
        filename += f"_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        # Čuvanje fajla u memoriji i slanje kao odgovor
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.read())
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        logger.info(f"Generisan Excel spisak učenika. Filteri: razred={student_class}, odeljenje={student_section}")
        return response
        
    except Exception as e:
        logger.error(f"Neočekivana greška pri izvozu učenika u Excel: {str(e)}")
        flash('Došlo je do greške pri generisanju Excel fajla.', 'danger')
        return redirect(url_for('students.student_list'))
